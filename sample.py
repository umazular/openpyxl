## coding: UTF-8
import os
import csv
import shutil
import openpyxl

ORGFILE = "sample.xlsm"
XLSFILE = "sampleoutput.xlsm"

# 画像ファイルを画像ワークシートに設定する.
def setImageData(wb, strIdentImage, strText, strFileImage):
    # 画像ファイルを読み込む.
    objImage = openpyxl.drawing.image.Image(strFileImage)
    if objImage is None: return
    # 画像ワークシートを取得する.
    ws = wb["images"]
    # 画像総数を加算する.
    intCol = ws["A1"].value
    if intCol is None: intCol = 0
    intCol = intCol + 1
    ws["A1"].value = intCol
    # 識別名をセルに設定する.
    strCellPos = ws.cell(row=2, column=intCol).coordinate
    ws[strCellPos].value = strIdentImage
    # 識別名をセルに設定する.
    strCellPos = ws.cell(row=3, column=intCol).coordinate
    ws[strCellPos].value = strText
    # 画像をセルに貼り付ける.
    strCellPos = ws.cell(row=4, column=intCol).coordinate
    ws.add_image(objImage, strCellPos)


# 雛形 Excel ブックファイルをコピーする.
shutil.copyfile(ORGFILE,XLSFILE)

# コピーした雛形 Excel ブックを開く
wb = openpyxl.load_workbook(filename=XLSFILE, read_only=False, keep_vba=True)

# データ CSV の数だけ処理を繰り返す(最大255).
for lngSheetCnt in range(1,255):
    # データファイル名・データワークシート名を設定する.
    strFileData = "data" + str(lngSheetCnt) + ".csv"
    strSheetData = "data" + str(lngSheetCnt)

    # データファイルが見つからない時、ループを終了する.
    if not os.path.isfile(strFileData): break

    # CSV ファイルを読み込む.
    objCsv = open("data" + str(lngSheetCnt) + ".csv")
    reader = csv.reader(objCsv)

    # シートリスト右端にデータワークシートを作成する.
    ws = wb.create_sheet(title=strSheetData)

    # CSV ファイル 1 行目を見出し情報として設定する.
    rownum = 1
    header = next(reader)
    ws.cell(rownum, 1).value = header[0]
    ws.cell(rownum, 2).value = header[1]

    # CSV ファイル 2 行目以降を明細情報として設定する.
    rownum = 2
    for rowdata in reader:
        ws.cell(rownum, 1).value = rowdata[0]
        ws.cell(rownum, 2).value = rowdata[1]
        ws.cell(rownum, 3).value = rowdata[2]
        rownum+=1

    # 画像ファイル名・画像識別子を設定する.
    strIdentImage = strSheetData + "!" + "QRCODE"
    strText = "This is QR Code"
    strFileImage = "qrcode" + str(lngSheetCnt) + ".png"

    # 画像ファイルを画像ワークシートに設定する.
    setImageData(wb, strIdentImage, strText, strFileImage)

# ワークブックを保存して終了する.
wb.save(XLSFILE)
wb.close

